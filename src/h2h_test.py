"""특정 팀이 각 상대 팀에게 보인 승률이, 그 매치업에서 '기대되는 승률'에
비춰봤을 때 통계적으로 유의하게 이례적인지를 이항검정(binomial test)으로 확인한다.

핵심 아이디어:
- 무승부는 제외하고(승/패만) 승률을 본다.
- 기대 승률은 Log5 공식(Bill James)으로 구한다: 내 팀과 상대팀 각각의
  '평소 실력' = 해당 상대(=서로)를 제외한 나머지 팀들을 상대로 한 승률
  (leave-one-out)을 구해 Log5에 대입한다. 자기 자신과의 대결 기록을
  각 팀의 기준(baseline)에 섞으면 안 되기 때문. 내 팀 자체 승률만 기준으로
  삼으면 "상대가 원래 강팀/약팀이라 그렇다"는 요인을 이례적 성적으로
  잘못 잡아낼 수 있다.
- H0: 이 상대에 대한 승률 = Log5 기대 승률
  H1(양측검정): 이 상대에 대한 승률은 Log5 기대 승률과 다르다.
- KBO는 팀당 상대 1팀과 16경기 내외만 치르기 때문에 표본이 매우 작다.
  즉 통계적 검정력이 낮아, 어느 정도 승률 차이가 나도 '우연'이라는 결론이
  나오기 쉽다는 점을 함께 봐야 한다.
"""
from pathlib import Path

import pandas as pd
from openpyxl.styles import Font, PatternFill
from rich.console import Console
from rich.table import Table
from scipy.stats import binomtest

ROOT = Path(__file__).resolve().parent.parent
RAW_DIR = ROOT / "data" / "raw"
OUTPUT_DIR = ROOT / "outputs"


def load_h2h(year: int) -> pd.DataFrame:
    return pd.read_csv(RAW_DIR / f"team_h2h_{year}.csv", encoding="utf-8-sig")


def _loo_win_rate(df: pd.DataFrame, team: str, exclude_opponent: str) -> float:
    """team의 승률(무승부 제외)을, exclude_opponent와의 맞대결만 뺀 나머지로 계산."""
    rows = df[(df["team"] == team) & (df["opponent"] != exclude_opponent)]
    win = rows["win"].sum()
    decisive = win + rows["loss"].sum()
    return win / decisive


def log5(win_rate_a: float, win_rate_b: float) -> float:
    """Bill James의 Log5: 두 팀의 평소 승률로 A가 B를 이길 기대 승률을 계산."""
    denom = win_rate_a + win_rate_b - 2 * win_rate_a * win_rate_b
    if denom <= 0:
        return win_rate_a
    return (win_rate_a - win_rate_a * win_rate_b) / denom


def analyze_team(df: pd.DataFrame, my_team: str) -> pd.DataFrame:
    mine = df[df["team"] == my_team].copy()
    mine["decisive"] = mine["win"] + mine["loss"]

    rows = []
    for _, r in mine.iterrows():
        opponent = r["opponent"]
        my_baseline = _loo_win_rate(df, my_team, opponent)
        opp_baseline = _loo_win_rate(df, opponent, my_team)
        expected_p = log5(my_baseline, opp_baseline)

        result = binomtest(int(r["win"]), int(r["decisive"]), expected_p, alternative="two-sided")

        rows.append(
            {
                "opponent": opponent,
                "record": f"{r['win']}-{r['loss']}-{r['draw']}",
                "win_rate_vs_opp": round(r["win"] / r["decisive"], 3),
                "opp_win_rate": round(opp_baseline, 3),
                "log5_expected": round(expected_p, 3),
                "diff": round(r["win"] / r["decisive"] - expected_p, 3),
                "p_value": round(result.pvalue, 4)
            }
        )

    out = pd.DataFrame(rows).sort_values("diff", ascending=True)
    return out


def print_table(result: pd.DataFrame, year: int, my_team: str, overall_win: int, overall_decisive: int) -> None:
    console = Console(width=110)
    console.print(
        f"\n[bold]{year}시즌 {my_team}[/bold] 전체 승률(무승부 제외): "
        f"{overall_win}/{overall_decisive} = {overall_win / overall_decisive:.3f}\n"
    )

    table = Table(show_lines=True)
    for col in result.columns:
        table.add_column(col, justify="right" if col != "opponent" and col != "record" else "left")

    for _, r in result.iterrows():
        significant = r["p_value"] < 0.05
        diff_style = "green" if r["diff"] > 0 else "red" if r["diff"] < 0 else "white"
        row_style = "bold" if significant else None
        table.add_row(
            str(r["opponent"]),
            str(r["record"]),
            f"{r['win_rate_vs_opp']:.3f}",
            f"{r['opp_win_rate']:.3f}",
            f"{r['log5_expected']:.3f}",
            f"[{diff_style}]{r['diff']:+.3f}[/{diff_style}]",
            f"{r['p_value']:.4f}{' *' if significant else ''}",
            style=row_style,
        )

    console.print(table)
    console.print(
        "\n[dim]주의: 팀당 상대 표본이 16경기 안팎이라 검정력이 낮습니다. "
        "p-value가 크다고 '문제 없다'가 아니라 '이 정도 표본으로는 우연과 구별이 안 된다'는 뜻입니다. "
        "(* = p < 0.05)[/dim]"
    )


def export_excel(result: pd.DataFrame, year: int, my_team: str) -> Path:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    out_path = OUTPUT_DIR / f"h2h_{year}_{my_team}.xlsx"

    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        result.to_excel(writer, index=False, sheet_name="h2h")
        ws = writer.sheets["h2h"]

        for col in ws.columns:
            width = max(len(str(cell.value)) for cell in col) + 2
            ws.column_dimensions[col[0].column_letter].width = width

        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        for cell in ws[1]:
            cell.font = Font(color="FFFFFF", bold=True)
            cell.fill = header_fill

        diff_col = result.columns.get_loc("diff") + 1
        p_col = result.columns.get_loc("p_value") + 1
        sig_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        for row in range(2, ws.max_row + 1):
            diff_cell = ws.cell(row=row, column=diff_col)
            diff_cell.font = Font(color="C00000" if diff_cell.value < 0 else "006100")
            if ws.cell(row=row, column=p_col).value < 0.05:
                for cell in ws[row]:
                    cell.fill = sig_fill

    return out_path


def main(year: int, my_team: str):
    df = load_h2h(year)
    result = analyze_team(df, my_team)

    overall = df[df["team"] == my_team]
    overall_win = int(overall["win"].sum())
    overall_decisive = int((overall["win"] + overall["loss"]).sum())

    print_table(result, year, my_team, overall_win, overall_decisive)
    out_path = export_excel(result, year, my_team)
    print(f"\n엑셀 저장 완료: {out_path}")


if __name__ == "__main__":
    main(year=2026, my_team="KIA")
